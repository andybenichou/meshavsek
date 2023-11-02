# -*- coding: utf-8 -*-

import csv
import os
import random
from copy import deepcopy

import pandas as pd

from collections import defaultdict
from datetime import datetime
from itertools import cycle

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from Guard import Guard
from GuardsList import GuardsList
from get_data import get_data


RANDOMNESS_LEVEL = 2
CRITICAL_DELAY = 9

# Define the week days
week_days = ['א', 'ב', 'ג', 'ד', 'ה', 'ו', 'שבת']

# List of guards
guards_list = ['יואל', 'ארד', 'ליאור', 'אבנר', 'משה', 'יונג', 'דורון',
               'אסרף', 'שגיא', 'אנדי', 'אנזו', 'דוד', 'דימנטמן', 'מטמוני',
               'דעאל', 'אגומס', 'ניסנוב', 'לואיס', 'דובר', 'כלפה', 'אלכסיי',
               'איתי כהן', 'עמיחי', 'לומיאנסקי', 'שמעון', 'דותן', 'קריספין',
               'רווה', 'דבוש', 'פיאצה', 'שראל', 'שרעבי', 'אסף', 'דימה',
               'שבצוב', 'נפמן', 'סדון', 'סיני', 'לוטם', 'אור',
               'בן', 'נח', 'לישי', 'מאור', 'רועי', 'משה החופל']

not_guarding = ['בן', 'נח', 'לישי', 'מאור', 'רועי', 'משה החופל']

northern_guards = ['לומיאנסקי', 'דובר', 'פיאצה', 'שראל', 'דבוש', 'ליאור',
                   'לישי', 'קריספין']

# List of missing guards each day
missing_guards = {
    'א': ['סדון', 'נפמן', 'לומיאנסקי', 'שגיא', 'אסרף'],
    'ב': ['שמעון', 'דימה', 'שבצוב', 'אור', 'ניסנוב', 'נפמן', 'דורון'],
    'ג': ['לואיס', 'ארד', 'קריספין', 'כלפה', 'אבנר', 'דעאל', 'לוטם', 'ניסנוב'],
    'ד': ['שרעבי', 'דוד', 'אנדי', 'אנזו', 'ניסנוב', 'יואל',
          'ליאור', 'סיני', 'לוטם'],
    'ה': ['אסף', 'פיאצה', 'רווה', 'דבוש', 'משה', 'שראל', 'ניסנוב', 'לישי'],
    'ו': ['אלכסיי', 'דותן', 'דובר', 'עמיחי', 'מטמוני', 'דימנטמן', 'ניסנוב',
          'יונג', 'שגיא'],
    'שבת': ['אלכסיי', 'דותן', 'דובר', 'עמיחי', 'מטמוני', 'דימנטמן', 'ניסנוב',
            'יונג', 'שגיא']
}

not_available_times_per_guard = {
    'לוטם': [{'start': {'day': 'ד', 'hour': 12},
              'end': {'day': 'ד', 'hour': 19}}],
    'דבוש': [{'start': {'day': 'ד', 'hour': 17},
              'end': {'day': 'ה', 'hour': 1}}],
    'פיאצה': [{'start': {'day': 'ד', 'hour': 17},
              'end': {'day': 'ה', 'hour': 1}}],
    'שראל': [{'start': {'day': 'ד', 'hour': 17},
              'end': {'day': 'ה', 'hour': 1}}],
    'מטמוני': [{'start': {'day': 'ד', 'hour': 17},
                'end': {'day': 'ה', 'hour': 1}}],
    'אסרף': [{'start': {'day': 'ד', 'hour': 17},
              'end': {'day': 'ה', 'hour': 1}}]
}


def get_next_week_day(day):
    day_i = week_days.index(day)

    return week_days[(day_i + 1) % 7]


def get_guards_list_obj(guards_list_prop):
    guards_list_obj = GuardsList()
    for guard in guards_list_prop:
        if guard not in guards_list_obj:
            not_available_times = list()
            for day in missing_guards:
                if guard in missing_guards[day]:
                    if guard in northern_guards:
                        time_obj = {
                            'start': {'day': day, 'hour': 6},
                            'end': {'day': get_next_week_day(day), 'hour': 16}
                        }
                    else:
                        time_obj = {
                            'start': {'day': day, 'hour': 9},
                            'end': {'day': get_next_week_day(day), 'hour': 12}
                        }

                    not_available_times.append(time_obj)

            if guard in not_available_times_per_guard:
                not_available_times.extend(not_available_times_per_guard[guard])

            guards_list_obj.append(Guard(guard,
                                         not_available_times=not_available_times))

    return guards_list_obj


# List of duos
duos = [('אנדי', 'דוד'), ('דימנטמן', 'מטמוני'), ('שבצוב', 'דימה'),
        ('עמיחי', 'איתי כהן'), ('דבוש', 'פיאצה'), ('שראל', 'שרעבי'),
        ('אלכסיי', 'לומיאנסקי'), ('סיני', 'לוטם'), ('דעאל', 'אגומס'),
        ('דובר', 'כלפה'), ('קריספין', 'רווה'), ('יואל', 'ארד'),
        ('נפמן', 'סדון')]

# Define the guard spots and their time slots
guard_spots = {
    'ש.ג.': ['0200-0500', '0500-0800', '0800-1100', '1100-1400',
             '1400-1700', '1700-2000', '2000-2300', '2300-0200'],
    'בטונדות': ['0200-0500', '0500-0800', '0800-1100', '1100-1400',
                '1400-1700',
                '1700-2000', '2000-2300', '2300-0200'],
    'פנטאוז': ['0200-0500', '0500-0800', '0800-1100', '1100-1400',
               '1400-1700', '1700-2000', '2000-2300', '2300-0200'],
    'פטרול': ['2200-0200', '0200-0600', '0600-1000'],
}

guards_number_per_spots = {
    'ש.ג.': 2,
    'בטונדות': 2,
    'פנטאוז': 2,
    'פטרול': 2,
}


def get_guards_slots(watch_list, days_prop):
    guard_slots = {guard: list() for guard in guards_list}

    for guard in guards_list:
        for day in watch_list:
            for time in watch_list[day]:
                for spot in watch_list[day][time]:
                    hour = int(time[:2])

                    if guard in watch_list[day][time][spot]:
                        guard_slots[guard].append((day, hour))

        if guard_slots[guard]:
            slot = guard_slots[guard][0]

            while type(slot) is tuple:
                guard_slots[guard].remove(slot)

                day = slot[0]
                start = slot[1]
                slot_obj = {
                    'start': (day, start),
                }

                days_cycle = cycle(days_prop)
                day_cycle = next(days_cycle)
                while day_cycle != day:
                    day_cycle = next(days_cycle)

                next_hour = start + 1
                next_day = day
                if next_hour == 24:
                    next_hour %= 24
                    next_day = next(days_cycle)

                while (next_day, next_hour) in guard_slots[guard]:
                    guard_slots[guard].remove((next_day, next_hour))
                    slot_obj['end'] = (next_day, next_hour + 1)
                    next_hour += 1
                    if next_hour == 24:
                        next_hour %= 24
                        next_day = next(days_cycle)

                guard_slots[guard].append(slot_obj)

                if guard_slots[guard]:
                    slot = guard_slots[guard][0]

    return guard_slots


def check_guards_slots_delays(watch_list, days_prop, need_print=False):
    guard_slots = get_guards_slots(watch_list, days_prop)
    bad_delays = list()
    too_good_delays = list()

    for guard in guard_slots:
        last_slot_end_day, last_slot_end_hour = None, None
        for slot in guard_slots[guard]:
            start_day, start_hour = slot['start']

            if last_slot_end_hour and last_slot_end_day:
                days_cycle = cycle(days_prop)
                curr_day = next(days_cycle)
                while curr_day != last_slot_end_day:
                    curr_day = next(days_cycle)

                while curr_day != start_day:
                    curr_day = next(days_cycle)
                    start_hour += 24

                delay = start_hour - last_slot_end_hour

                if delay <= CRITICAL_DELAY:
                    bad_delays.append({
                        'guard': guard,
                        'delay': delay,
                        'beggining': {
                            'day': start_day,
                            'hour': slot["start"][1]
                        }
                    })
                elif CRITICAL_DELAY + 3 < delay <= CRITICAL_DELAY + 12:
                    too_good_delays.append({
                        'guard': guard,
                        'delay': delay,
                        'beggining': {
                            'day': start_day,
                            'hour': slot["start"][1]
                        }
                    })

            last_slot_end_day, last_slot_end_hour = slot['end']

    if need_print:
        # Map each day to its numeric value
        day_to_num = {day: num for num, day in enumerate(week_days)}

        # Sort the data
        bad_delays = sorted(bad_delays, key=lambda x: (day_to_num[x['beggining']['day']], x['beggining']['hour']))
        too_good_delays = sorted(too_good_delays, key=lambda x: (day_to_num[x['beggining']['day']], x['beggining']['hour']))

        for b_d in bad_delays:
            print(
                f'{b_d["guard"]} יש לו רק {b_d["delay"]} שעות מנוחה לפני המשמרת ביום {b_d["beggining"]["day"]} בשעה {b_d["beggining"]["hour"]}')

        if bad_delays:
            print()

        for g_d in too_good_delays:
            print(
                f'{g_d["guard"]} יש לו {g_d["delay"]} שעות מנוחה לפני המשמרת ביום {g_d["beggining"]["day"]} בשעה {g_d["beggining"]["hour"]}')

    return len(bad_delays)


def get_prec_day(day_prop, days_prop):
    day_i = days_prop.index(day_prop)
    prec_day = None

    if day_i > 0:
        prec_day = days_prop[day_i - 1]

    return prec_day


# Helper function to check if a guard is available
def is_guard_available(watch_list, guard: Guard, day_prop, time_prop, days_prop,
                       delays_prop=None):
    if guard.name in not_guarding:
        return False

    hour = int(time_prop[:2])

    if guard.is_missing(day_prop, hour):
        return False

    for rest_delay in (delays_prop if delays_prop else [0, 3, 6, 9]):
        updated_hour = hour - rest_delay
        updated_day = day_prop

        if updated_hour < 0:
            updated_day = get_prec_day(day_prop, days_prop)
            updated_hour += 24

        if not updated_day:
            break

        # Check if the guard already in another spot
        for p in guard_spots:
            for t in guard_spots[p]:
                beginning_str, g_end_str = t.split('-')
                beginning, g_end = map(int, [beginning_str[:2], g_end_str[:2]])

                if g_end < beginning:
                    g_end += 24

                    if beginning <= updated_hour + 24 < g_end:
                        updated_hour += 24

                slot_d = updated_day

                if beginning <= updated_hour < g_end:
                    if updated_hour >= 24:
                        slot_d = get_prec_day(updated_day, days_prop)

                        if not slot_d:
                            break

                    if guard in watch_list[slot_d][beginning_str][p]:
                        return False

    return True


# Align guards cycle to the next available guard
def align_guards_cycle(watch_list, guard_cycle_prop,
                       guards_list_obj: GuardsList,
                       day_prop, time_prop, days_prop,
                       currently_missing_guards):
    guard = guards_list_obj.get_current_guard()

    if not guard:
        guard = next(guard_cycle_prop)
        guards_list_obj.set_current_guard(guard)

    while not is_guard_available(watch_list, guard, day_prop,
                                 time_prop, days_prop):

        if guard.is_missing(day_prop, time_prop):
            currently_missing_guards.append(guard)

        guard = next(guard_cycle_prop)
        guards_list_obj.set_current_guard(guard)

    return guard


# Helper function to get the next available guard
def get_next_available_guard(guards_list_obj: GuardsList,
                             watch_list, guard_cycle_prop, day_prop,
                             time_prop, days_prop, currently_missing_guards,
                             spot, guards: GuardsList = None, no_duo=False):
    # if time_prop == '0500' \
    #         and spot == 'פנטאוז' \
    #         and is_guard_available(watch_list, 'משה החופל',
    #                                day_prop, time_prop,
    #                                days_prop) \
    #         and 'משה החופל' not in guards:
    #     return 'משה החופל', None

    curr_guard = align_guards_cycle(watch_list, guard_cycle_prop,
                                    guards_list_obj,
                                    day_prop, time_prop, days_prop,
                                    currently_missing_guards)
    index = guards_list_obj.index(curr_guard)
    buff_cycle = cycle(guards_list_obj[index:] + guards_list_obj[:index])

    while True:
        random_guards = list()
        while len(random_guards) != RANDOMNESS_LEVEL:
            guard = None
            for g in currently_missing_guards:
                if is_guard_available(watch_list, g,
                                      day_prop, time_prop, days_prop):
                    currently_missing_guards.remove(g)
                    break

            if not guard:
                guard = next(buff_cycle)

            if is_guard_available(watch_list, guard, day_prop,
                                  time_prop, days_prop) \
                    and guard not in guards:
                random_guards.append(guard)

        random.shuffle(random_guards)

        for guard in random_guards:
            duo = next((duo for duo in duos if guard in duo), None)

            if duo:
                partner = duo[0] if duo[1] == guard else duo[1]
                partner_obj = guards_list_obj.find(partner)

                if partner_obj.is_missing(day_prop, time_prop):
                    return guard, None

                if not is_guard_available(watch_list, partner_obj, day_prop,
                                          time_prop, days_prop,
                                          delays_prop=[0, 3, 6]):
                    return guard, None

                if no_duo:
                    continue

                elif is_guard_available(watch_list, partner_obj, day_prop,
                                        time_prop, days_prop):
                    return guard, partner

            else:
                return guard, None


def get_today_day_of_week():
    # Get the current date
    today = datetime.today()

    # Get the name of the day of the week
    day_name = today.strftime("%A")

    return day_name


def get_days(watch_list):
    user_input = input("How many days do you need to schedule? ")

    while True:
        if user_input.isdigit():
            days_num = int(user_input)
            break
        else:
            user_input = input("Please enter a valid integer. ")

    days_list = list(
        watch_list.keys()) if watch_list.keys() else get_today_day_of_week()
    days_cycle = cycle(week_days)

    day = next(days_cycle)
    while days_list[-1] != day:
        day = next(days_cycle)

    for _ in range(days_num):
        day = next(days_cycle)

        if day in days_list:
            break

        days_list.append(day)

    return days_list


def get_already_filled_guard_slot(watch_list, day, time, hour, spot,
                                  days_prop):
    guards = watch_list[day][time][spot] if watch_list[day][time][
        spot] else list()
    fill_guard_spot = False

    # Spot already filled
    if len(guards) == guards_number_per_spots[spot]:
        return guards, fill_guard_spot

    for slot in guard_spots[spot]:
        # Fill the actual hour with the slot guards if there are
        # already in one of the slot hour
        start_str, end_str = slot.split('-')
        start, end = int(start_str[:2]), int(end_str[:2])

        if end < start:
            end += 24

            if start <= hour + 24 < end:
                hour += 24

        if start <= hour < end:
            fill_guard_spot = True

        if start < hour < end:
            slot_day = day
            if hour == 24:
                day_index = days_prop.index(day)

                if day_index > 0:
                    slot_day = days_prop[day_index - 1]

            guards = watch_list[slot_day][f'{((hour - 1) % 24):02d}00'][spot]

    return guards, fill_guard_spot


def get_guards(guards_list_obj: GuardsList, watch_list, guard_cycle, day,
               time, hour, spot, days_prop, currently_missing_guards):
    guards, fill_guard_spot = \
        get_already_filled_guard_slot(watch_list,
                                      day, time, hour, spot,
                                      days_prop)

    guards = GuardsList(guards)

    if len(guards) == guards_number_per_spots[spot]:
        return guards

    if fill_guard_spot:
        guard1, guard2 = get_next_available_guard(guards_list_obj,
                                                  watch_list,
                                                  guard_cycle,
                                                  day,
                                                  time,
                                                  days_prop,
                                                  currently_missing_guards,
                                                  spot,
                                                  guards=guards)

        for g in [guard1, guard2]:
            if len(guards) != guards_number_per_spots[spot] and g:
                guards.append(g)

        if len(guards) == 1 and guards_number_per_spots[spot] == 2:
            guard2 = get_next_available_guard(guards_list_obj,
                                              watch_list,
                                              guard_cycle,
                                              day,
                                              time,
                                              days_prop,
                                              currently_missing_guards,
                                              spot,
                                              guards=guards,
                                              no_duo=True)[0]
            guards.append(guard2)
        guards.sort()

    return guards


def get_watch_list_data(guards_list_obj: GuardsList, watch_list, days_prop,
                        first_hour_prop):
    currently_missing_guards = list()

    # Assign guards to each slot
    guard_cycle = cycle(guards_list_obj)
    for day in days_prop:
        for time in sorted(set([f'{hour:02d}00' for hour in range(24)])):
            hour = int(time[:2])

            if day == days_prop[0] and hour < first_hour_prop:
                continue

            # Remove unnecessary start and end of planning
            if (days_prop.index(day) == 0 and hour < 2) or \
                    (days_prop.index(day) == len(days_prop) - 1
                     and hour >= 20):
                continue

            for spot in guard_spots:
                guards = get_guards(guards_list_obj, watch_list, guard_cycle,
                                    day, time, hour, spot, days_prop,
                                    currently_missing_guards)

                if len(guards) == guards_number_per_spots[spot] \
                        and not watch_list[day][time][spot]:
                    watch_list[day][time][spot].extend(guards)

    return watch_list


def export_to_CSV(watch_list, days_prop):
    with open('watch_list.csv', 'w', newline='') as csvfile:
        csvwriter = csv.writer(csvfile)
        header = ['יום', 'שעה'] + list(guard_spots.keys())
        csvwriter.writerow(header)
        for day in days_prop:
            for hour in range(24):
                time = f'{hour:02d}00'
                row = [day, time]

                for spot in guard_spots.keys():
                    guards_str = '\n'.join(watch_list[day][time][spot]) \
                        if watch_list[day][time][spot] else ' '
                    row.append(guards_str)
                csvwriter.writerow(row)


def get_excel_data_frame(watch_list, days_prop):
    columns = ['יום', 'שעה'] + list(guard_spots.keys())
    data = list()

    for day in days_prop:
        for hour in range(24):
            if (days_prop.index(day) == 0 and hour < 2) or (
                    days_prop.index(day) == len(days_prop) - 1 and hour >= 20):
                continue

            time = f'{hour:02d}00'
            time_for_excel = f'{hour:02d}:00'  # formatted for Excel
            row = [day, time_for_excel]

            for spot in guard_spots.keys():
                guards_str = '\n'.join(
                    [g.name for g in watch_list[day][time][spot]]) \
                    if watch_list[day][time][spot] else ' '
                row.append(guards_str)

            is_row_empty = True
            for guards_str in row[2:]:
                if len(guards_str) > 1:
                    is_row_empty = False

            if not is_row_empty:
                data.append(row)

    return pd.DataFrame(data, columns=columns)


def merge_excel_cells(df, worksheet):
    # Merge adjacent cells with the same content
    for col_num in range(1, len(df.columns) + 2):
        start_row = 1
        start_content = None

        for row_num in range(1, len(df) + 2):
            content = worksheet.cell(row=row_num, column=col_num).value

            if start_content is None:
                start_content = content

            elif content != start_content or row_num == len(df) + 1:
                if start_row < row_num - 1:
                    end_row = row_num - 1 if row_num < len(df) + 1 else row_num
                    worksheet.merge_cells(start_row=start_row,
                                          start_column=col_num,
                                          end_row=end_row, end_column=col_num)

                if row_num < len(df) + 1:
                    start_row = row_num
                    start_content = content


def adjust_columns_and_rows(worksheet):
    # Adjust the column width according to content, wrap text, and center align
    # text
    for col in worksheet.columns:
        max_length = 0
        column = [cell for cell in col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

                # Set wrap text and center alignment
                cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                           vertical='center')
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[
            column[0].column_letter].width = adjusted_width


def format_excel(df, worksheet):
    merge_excel_cells(df, worksheet)
    adjust_columns_and_rows(worksheet)


def export_to_excel(file_name, watch_list, days_prop):
    df = get_excel_data_frame(watch_list, days_prop)

    # Save to Excel
    excel_path = f'{file_name}.xlsx'
    df.to_excel(excel_path, index=False, sheet_name='שבצ״כ')

    workbook = load_workbook(filename=excel_path)
    worksheet = workbook.active

    format_excel(df, worksheet)

    workbook.save(filename=excel_path)


if __name__ == '__main__':
    # Initialize the watch list
    wl = defaultdict(lambda: defaultdict(lambda: defaultdict(GuardsList)))

    old_file_name = 'old'
    src_dir = os.path.dirname(os.path.abspath(__file__))
    old_dir = os.path.join(src_dir, f'{old_file_name}.xlsx')

    guards_list_object = get_guards_list_obj(guards_list)
    if os.path.exists(old_dir):
        wl = get_data(old_file_name, wl, guards_list_object, guard_spots)

    days = get_days(wl)

    i = 0
    min_delays = 0
    min_i = 0
    # Initialize the watch list
    guards_list_object = get_guards_list_obj(guards_list)
    wl = defaultdict(lambda: defaultdict(lambda: defaultdict(GuardsList)))
    wls = list()
    while i < 10:
        print(i)
        wl = defaultdict(lambda: defaultdict(lambda: defaultdict(GuardsList)))
        if os.path.exists(old_dir):
            wl = get_data(old_file_name, wl, guards_list_object, guard_spots,
                          print_missing_names=False)

        for d in days:
            if d not in wl.keys():
                wl[d] = defaultdict(lambda: defaultdict(GuardsList))

        # Get first hour of first day:
        first_hour = None
        for d in days:
            for str_hour in sorted(set([f'{hour:02d}00' for hour in range(24)])):
                if str_hour in wl[d].keys() and not first_hour:
                    first_hour = int(str_hour[:2])

        buff_wl = deepcopy(wl)

        buff_wl = get_watch_list_data(guards_list_object, buff_wl, days,
                                      first_hour)

        delays = check_guards_slots_delays(buff_wl, days)

        wls.append(buff_wl)

        if not min_delays or delays < min_delays:
            min_delays = delays
            min_i = i
        i += 1

    export_to_excel('watch_list', wls[min_i], days)

    check_guards_slots_delays(wls[min_i], days, need_print=True)

    print('\nShivsakta!')
