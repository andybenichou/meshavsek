import os
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook

from config import DAY_COLUMN_NAME, AVAILABLE_GUARDS_COLUMN_NAME, \
    HOUR_COLUMN_NAME, AVAILABLE_GUARDS_SHEET_NAME
from src.models.GuardsList import GuardsList
from src.services.export import parse_date, is_row_empty, format_excel


def format_to_excel_data_frame(available_guards_per_date):
    columns = [DAY_COLUMN_NAME, HOUR_COLUMN_NAME, AVAILABLE_GUARDS_COLUMN_NAME]
    data = list()

    for date in available_guards_per_date:
        time_for_excel = f'{date.hour:02d}:00'  # formatted for Excel
        row = [parse_date(date), time_for_excel]

        guards_str = ''
        for i, g in enumerate(available_guards_per_date[date]):
            if i == 0:
                guards_str += str(g)

            elif 0 < i < len(available_guards_per_date[date]) and i % 5 == 0:
                guards_str += '\n' + str(g)

            else:
                guards_str += ', ' + str(g)

        row.append(guards_str)

        if not is_row_empty(row):
            data.append(row)

    return pd.DataFrame(data, columns=columns)


def export_to_excel(file_name, available_guards_per_date):
    df = format_to_excel_data_frame(available_guards_per_date)

    # Define the Excel file path
    excel_path = f'data/output/{file_name}.xlsx'

    # Check if the Excel file exists
    if os.path.exists(excel_path):
        # Load the existing workbook
        workbook = load_workbook(excel_path)

        # Check if the sheet exists
        if AVAILABLE_GUARDS_SHEET_NAME in workbook.sheetnames:
            # Remove existing sheet
            del workbook[AVAILABLE_GUARDS_SHEET_NAME]

        # Add a new sheet with the new data
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        writer.book = workbook
        df.to_excel(writer, index=False, sheet_name=AVAILABLE_GUARDS_SHEET_NAME)
        writer.save()
    else:
        # Create a new workbook and save the data
        df.to_excel(excel_path, index=False, sheet_name=AVAILABLE_GUARDS_SHEET_NAME)

    workbook = load_workbook(filename=excel_path)
    worksheet = workbook[AVAILABLE_GUARDS_SHEET_NAME]
    format_excel(df, worksheet)
    workbook.save(filename=excel_path)


def get_available_guards_per_date(wl, guards_list, file_name, backward_delay=0,
                                  forward_delay=0):
    available_guards_per_date = defaultdict(GuardsList)
    for index, date in enumerate(wl):
        if index < backward_delay or index >= len(wl) - forward_delay:
            continue

        for guard in guards_list:
            if guard.is_available(wl, date, spot=None,
                                  delays_prop=list(range(-backward_delay,
                                                         forward_delay + 1, 3)),
                                  not_missing_delay=forward_delay) \
                    and guard not in available_guards_per_date[date]:
                available_guards_per_date[date].append(guard)

    export_to_excel(file_name, available_guards_per_date)

    return available_guards_per_date
