import csv
import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.exceptions import IllegalCharacterError

from config import DAY_COLUMN_NAME, HOUR_COLUMN_NAME
from guards_config import KITOT_KONENOUT_PROPS, TORANOUT_PROPS


def export_to_CSV(watch_list, guard_spots):
    with open('watch_list.csv', 'w', newline='') as csvfile:
        csvwriter = csv.writer(csvfile)
        header = ['יום', 'שעה'] + list(guard_spots.keys())
        csvwriter.writerow(header)
        for date in watch_list:
            row = [date]

            for spot in guard_spots.keys():
                guards_str = '\n'.join(watch_list[date][spot]) \
                    if watch_list[date][spot] else ' '
                row.append(guards_str)
            csvwriter.writerow(row)


def is_row_empty(row):
    is_empty = True
    for guards_str in row[2:]:
        if len(guards_str) > 1:
            is_empty = False

    return is_empty


def parse_date(date: datetime):
    return date.date()


def get_row(watch_list, guard_spots, date, duty_room_per_day, kitot_konenout):
    time_for_excel = f'{date.hour:02d}:00'  # formatted for Excel
    row = [parse_date(date), time_for_excel]

    for spot in guard_spots:
        guards_str = '\n'.join(
            [str(g) for g in watch_list[date][spot]]) \
            if watch_list[date][spot] else ' '

        row.append(guards_str)

    for spot in [KITOT_KONENOUT_PROPS['column_name'], TORANOUT_PROPS['column_name']]:
        if KITOT_KONENOUT_PROPS['is_needed']:
            if spot == KITOT_KONENOUT_PROPS['column_name'] \
                    and date in kitot_konenout \
                    and kitot_konenout[date] is not None:
                if kitot_konenout[date]:
                    row.append(f'{kitot_konenout[date]} חדר')
                else:
                    row.append(' ')
        if TORANOUT_PROPS['is_needed']:
            if spot == TORANOUT_PROPS['column_name'] \
                    and date in duty_room_per_day:
                if duty_room_per_day[date]:
                    row.append(f'{duty_room_per_day[date].number} חדר')
                else:
                    row.append(' ')

    return row


def get_limit_dates(watch_list, guard_spots, duty_room_per_day,
                    kitot_konenout):
    first_date = None
    last_date = None
    for date in watch_list:
        row = get_row(watch_list, guard_spots, date, duty_room_per_day,
                      kitot_konenout)

        if not is_row_empty(row):
            last_date = date

            if not first_date:
                first_date = date

    return first_date, last_date


def get_excel_data_frame(watch_list, guard_spots, duty_room_per_day,
                         kitot_konenout):
    first_date, last_date = get_limit_dates(watch_list, guard_spots,
                                            duty_room_per_day,
                                            kitot_konenout)
    columns = [DAY_COLUMN_NAME, HOUR_COLUMN_NAME]

    for spot in guard_spots:
        columns.append(spot.name)

    if KITOT_KONENOUT_PROPS['is_needed']:
        columns += KITOT_KONENOUT_PROPS['column_name']

    if TORANOUT_PROPS['is_needed']:
        columns += TORANOUT_PROPS['column_name']

    data = list()

    for date in watch_list:
        row = get_row(watch_list, guard_spots, date, duty_room_per_day,
                      kitot_konenout)

        if first_date <= date <= last_date:
            data.append(row)

    return remove_empty_columns(pd.DataFrame(data, columns=columns))


def merge_excel_cells(df, worksheet):
    # Merge adjacent cells with the same content
    for col_num in range(1, len(df.columns) + 1):
        start_row = 1
        start_content = None

        for row_num in range(1, len(df) + 2):
            content = worksheet.cell(row=row_num, column=col_num).value

            if start_content is None:
                start_content = content

            elif content != start_content:
                if start_row < row_num - 1:
                    end_row = row_num - 1
                    worksheet.merge_cells(start_row=start_row,
                                          start_column=col_num,
                                          end_row=end_row, end_column=col_num)

                if row_num < len(df) + 1:
                    start_row = row_num
                    start_content = content

            elif row_num == len(df) + 1 and content == start_content:
                if start_row < row_num:
                    end_row = row_num
                    worksheet.merge_cells(start_row=start_row,
                                          start_column=col_num,
                                          end_row=end_row, end_column=col_num)


def is_column_empty(worksheet, column_index):
    # Iterate through each cell in the column
    for row in worksheet.iter_rows(min_col=column_index, max_col=column_index):
        cell = row[0]  # Since we're looking at one column, there's only one cell per row tuple
        if cell.value is not None:
            return False  # Found a cell with data, so the column is not empty
    return True  # All cells in the column are empty


def adjust_columns_and_rows(worksheet):
    # Define font, alignment, border, and fill styles
    bold_font = Font(bold=True, size=12)
    center_aligned_text = Alignment(horizontal='center', vertical='center',
                                    text_rotation=0, wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    grey_fill = PatternFill(start_color='f2f2f2',  # Light grey color
                            end_color='f2f2f2',
                            fill_type='solid')

    # Adjusting column widths
    for col in worksheet.columns:
        max_length = 0
        column = [cell for cell in col]
        for cell in column:
            try:
                # Check if the cell has a value and it's not just whitespace
                if cell.value and (not isinstance(cell.value, str) or cell.value.strip()):
                    # Find the longest line in the cell
                    longest_line_length = \
                        max(len(line) for line in str(cell.value).split('\n'))
                    max_length = max(max_length, longest_line_length)
                else:
                    # Apply grey fill only to empty or whitespace-only cells
                    cell.fill = grey_fill

                # Apply styles to all cells
                cell.font = bold_font
                cell.alignment = center_aligned_text
                cell.border = thin_border
            except IllegalCharacterError:
                pass

        adjusted_width = (max_length + 2)  # Adjust the width
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Constants for height calculation
    line_height = 30  # Adjust this value as needed
    padding_height = 4  # Additional height to add as padding

    # Adjusting row heights
    for index, row in enumerate(worksheet.iter_rows()):
        if index < 1:
            continue

        max_line_count = 1
        row_index = row[0].row  # Get the row index from the first cell in the row
        for cell in row:
            try:
                if cell.value and isinstance(cell.value, str) and '\n' in cell.value:
                    line_count = cell.value.count('\n') + 1
                    if line_count > max_line_count:
                        max_line_count = line_count
            except IllegalCharacterError:
                pass
        worksheet.row_dimensions[row_index].height = (max_line_count * line_height) + padding_height

    # Adjusting for merged cells
    for merged_range in worksheet.merged_cells.ranges:
        try:
            first_cell = worksheet[merged_range.min_row][merged_range.min_col - 1]
            if first_cell.value and isinstance(first_cell.value, str) and '\n' in first_cell.value:
                line_count = first_cell.value.count('\n') + 1
                total_height = (line_count * line_height) + padding_height  # Calculate total height

                # Number of rows in the merged cell
                num_rows = merged_range.max_row - merged_range.min_row + 1

                # Distribute the height evenly across the rows
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    worksheet.row_dimensions[row].height = total_height / num_rows
        except IllegalCharacterError:
            pass


def remove_empty_columns(df):
    columns_to_remove = []

    for column_name in df.columns:
        if df[column_name].dtype == 'object':  # Check if the column contains object (usually string) data
            is_empty = all(df[column_name].astype(str).str.strip().eq(''))  # Convert to string and then check if all string values are empty

            if is_empty:
                columns_to_remove.append(column_name)

    # Remove the empty columns from the DataFrame
    df.drop(columns=columns_to_remove, inplace=True)
    return df


def format_excel(df, worksheet):
    merge_excel_cells(df, worksheet)
    adjust_columns_and_rows(worksheet)


def export_to_excel(file_name, watch_list, guard_spots, duty_room_per_day,
                    kitot_konenout):
    df = get_excel_data_frame(watch_list, guard_spots, duty_room_per_day,
                              kitot_konenout)

    # Save to Excel
    excel_path = f'data/output/{file_name}.xlsx'
    ROOT_DIR = os.environ.get('ROOT_DIR')
    full_path = os.path.join(ROOT_DIR, excel_path)

    df.to_excel(full_path, index=False, sheet_name='שבצ״כ')

    workbook = load_workbook(filename=full_path)
    worksheet = workbook.active

    format_excel(df, worksheet)

    workbook.save(filename=excel_path)
